"""
ServerHardenPro — Generador de Reportes PDF / Excel
"""

import datetime
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paleta
C_BG     = colors.HexColor('#080c10')
C_BG2    = colors.HexColor('#0d1117')
C_CYAN   = colors.HexColor('#00e5ff')
C_GREEN  = colors.HexColor('#39ff6e')
C_RED    = colors.HexColor('#ff3b5c')
C_YELLOW = colors.HexColor('#ffd600')
C_TEXT   = colors.HexColor('#cdd9e5')
C_MUTED  = colors.HexColor('#4a6274')
C_WHITE  = colors.white

def _score_color_rl(score):
    if score >= 80: return C_GREEN
    if score >= 60: return C_YELLOW
    return C_RED

def generate_pdf(data: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    story  = []

    # Estilos
    title_style = ParagraphStyle('T', parent=styles['Normal'],
        fontSize=20, fontName='Helvetica-Bold',
        textColor=C_WHITE, leading=26)
    sub_style = ParagraphStyle('S', parent=styles['Normal'],
        fontSize=9, fontName='Courier',
        textColor=C_CYAN, letterSpacing=2)
    section_style = ParagraphStyle('Sec', parent=styles['Normal'],
        fontSize=11, fontName='Helvetica-Bold',
        textColor=C_CYAN, spaceBefore=12, spaceAfter=6)
    body_style = ParagraphStyle('B', parent=styles['Normal'],
        fontSize=9, fontName='Helvetica',
        textColor=C_TEXT, leading=13)
    muted_style = ParagraphStyle('M', parent=styles['Normal'],
        fontSize=8, fontName='Courier', textColor=C_MUTED)

    srv  = data.get('server', {})
    summ = data.get('summary', {})
    cats = data.get('category_scores', {})
    chks = data.get('checks', [])
    score = summ.get('score', 0)

    # HEADER — separado en dos filas para evitar solapamiento
    header_data = [[
        Paragraph('<b>ServerHardenPro</b>', ParagraphStyle('H',
            parent=styles['Normal'], fontSize=22, fontName='Helvetica-Bold',
            textColor=C_WHITE)),
        Paragraph(f'<b>{score}%</b>', ParagraphStyle('SC',
            parent=styles['Normal'], fontSize=26, fontName='Helvetica-Bold',
            textColor=_score_color_rl(score), alignment=TA_RIGHT))
    ]]
    header_table = Table(header_data, colWidths=[11*cm, 6*cm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), C_BG2),
        ('TOPPADDING',    (0,0), (-1,-1), 18),
        ('BOTTOMPADDING', (0,0), (-1,-1), 18),
        ('LEFTPADDING',   (0,0), (-1,-1), 18),
        ('RIGHTPADDING',  (0,0), (-1,-1), 18),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW',     (0,0), (-1,0), 2, C_CYAN),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6))
    story.append(Paragraph('// REPORTE DE AUDITORIA DE HARDENING', sub_style))
    story.append(HRFlowable(width='100%', thickness=1, color=C_CYAN, spaceAfter=12))

    # INFO DEL SERVIDOR
    story.append(Paragraph('INFORMACION DEL SERVIDOR', section_style))
    info_data = [
        ['Hostname',     srv.get('hostname', 'N/A')],
        ['IP',           srv.get('ip', 'N/A')],
        ['Sistema Op.',  srv.get('os', 'N/A')],
        ['Plataforma',   srv.get('platform', 'linux').upper()],
        ['Fecha Audit.', str(summ.get('audit_date', datetime.datetime.now().isoformat()))[:19]],
    ]
    info_table = Table(info_data, colWidths=[4*cm, 13*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (0,-1), C_BG2),
        ('BACKGROUND',    (1,0), (1,-1), C_BG),
        ('TEXTCOLOR',     (0,0), (0,-1), C_CYAN),
        ('TEXTCOLOR',     (1,0), (1,-1), C_TEXT),
        ('FONTNAME',      (0,0), (0,-1), 'Courier-Bold'),
        ('FONTNAME',      (1,0), (1,-1), 'Helvetica'),
        ('FONTSIZE',      (0,0), (-1,-1), 9),
        ('TOPPADDING',    (0,0), (-1,-1), 7),
        ('BOTTOMPADDING', (0,0), (-1,-1), 7),
        ('LEFTPADDING',   (0,0), (-1,-1), 10),
        ('GRID',          (0,0), (-1,-1), 0.5, C_MUTED),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 12))

    # RESUMEN
    story.append(Paragraph('RESUMEN DE RESULTADOS', section_style))
    summ_data = [
        ['TOTAL', 'PASS', 'FAIL', 'WARN', 'SCORE'],
        [
            str(summ.get('total', 0)),
            str(summ.get('pass', 0)),
            str(summ.get('fail', 0)),
            str(summ.get('warn', 0)),
            f"{score}%"
        ]
    ]
    summ_table = Table(summ_data, colWidths=[3.4*cm]*5)
    summ_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), C_BG2),
        ('TEXTCOLOR',     (0,0), (-1,0), C_CYAN),
        ('FONTNAME',      (0,0), (-1,0), 'Courier-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 10),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('BACKGROUND',    (1,1), (1,1), colors.HexColor('#0d2b1a')),
        ('TEXTCOLOR',     (1,1), (1,1), C_GREEN),
        ('FONTNAME',      (1,1), (1,1), 'Helvetica-Bold'),
        ('BACKGROUND',    (2,1), (2,1), colors.HexColor('#2b0d15')),
        ('TEXTCOLOR',     (2,1), (2,1), C_RED),
        ('FONTNAME',      (2,1), (2,1), 'Helvetica-Bold'),
        ('BACKGROUND',    (3,1), (3,1), colors.HexColor('#2b2500')),
        ('TEXTCOLOR',     (3,1), (3,1), C_YELLOW),
        ('FONTNAME',      (3,1), (3,1), 'Helvetica-Bold'),
        ('TEXTCOLOR',     (4,1), (4,1), _score_color_rl(score)),
        ('FONTNAME',      (4,1), (4,1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,1), (-1,1), 16),
        ('GRID',          (0,0), (-1,-1), 0.5, C_MUTED),
    ]))
    story.append(summ_table)
    story.append(Spacer(1, 12))

    # CATEGORIAS
    if cats:
        story.append(Paragraph('CUMPLIMIENTO POR CATEGORIA', section_style))
        cat_rows = [['Categoria', 'Score', 'Estado']]
        for cat, pct in cats.items():
            status = 'OK' if pct >= 80 else 'ADVERTENCIA' if pct >= 60 else 'CRITICO'
            cat_rows.append([cat, f'{pct}%', status])
        cat_table = Table(cat_rows, colWidths=[7*cm, 3*cm, 7*cm])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), C_BG2),
            ('TEXTCOLOR',     (0,0), (-1,0), C_CYAN),
            ('FONTNAME',      (0,0), (-1,0), 'Courier-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [C_BG, C_BG2]),
            ('TEXTCOLOR',     (0,1), (-1,-1), C_TEXT),
            ('TOPPADDING',    (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING',   (0,0), (-1,-1), 10),
            ('GRID',          (0,0), (-1,-1), 0.5, C_MUTED),
        ]))
        story.append(cat_table)
        story.append(Spacer(1, 12))

    # CHECKS DETALLADOS
    story.append(Paragraph('CHECKS DETALLADOS', section_style))
    chk_rows = [['Estado', 'Categoria', 'Verificacion', 'Sev.']]
    for c in chks:
        status_txt = c['status']
        chk_rows.append([
            Paragraph(status_txt, body_style),
            Paragraph(c['category'], muted_style),
            Paragraph(f"<b>{c['name']}</b><br/><font size='7' color='#4a6274'>{c['description']}</font>", body_style),
            Paragraph(c['severity'], muted_style),
        ])
    chk_table = Table(chk_rows, colWidths=[2*cm, 3*cm, 10*cm, 2*cm])
    chk_table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), C_BG2),
        ('TEXTCOLOR',     (0,0), (-1,0), C_CYAN),
        ('FONTNAME',      (0,0), (-1,0), 'Courier-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 8),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [C_BG, C_BG2]),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('GRID',          (0,0), (-1,-1), 0.3, C_MUTED),
    ]))
    story.append(chk_table)

    # FOOTER
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width='100%', thickness=1, color=C_MUTED))
    story.append(Paragraph(
        f'Generado por ServerHardenPro v0.5 · {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}',
        ParagraphStyle('Footer', parent=styles['Normal'],
            fontSize=7, fontName='Courier',
            textColor=C_MUTED, alignment=TA_CENTER, spaceBefore=6)
    ))

    doc.build(story)
    return buffer.getvalue()


def generate_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()

    srv  = data.get('server', {})
    summ = data.get('summary', {})
    cats = data.get('category_scores', {})
    chks = data.get('checks', [])

    BG_DARK = '080c10'; BG_MED = '0d1117'
    CYAN = '00e5ff'; GREEN = '39ff6e'; RED = 'ff3b5c'
    YELLOW = 'ffd600'; TEXT = 'cdd9e5'; MUTED = '4a6274'

    def fill(h): return PatternFill('solid', fgColor=h)
    def bfont(c='FFFFFF', s=10): return Font(bold=True, color=c, size=s, name='Consolas')
    def rfont(c=TEXT, s=9): return Font(color=c, size=s, name='Calibri')
    def tborder():
        s = Side(style='thin', color=MUTED)
        return Border(left=s, right=s, top=s, bottom=s)
    def ctr(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def lft(): return Alignment(horizontal='left', vertical='center', wrap_text=True)

    # HOJA 1: RESUMEN
    ws1 = wb.active; ws1.title = 'Resumen'
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 35

    ws1.merge_cells('A1:B1')
    ws1['A1'] = 'ServerHardenPro — Reporte de Auditoria'
    ws1['A1'].font = Font(bold=True, color=CYAN, size=14, name='Consolas')
    ws1['A1'].fill = fill(BG_MED)
    ws1['A1'].alignment = ctr()
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells('A2:B2')
    ws1['A2'] = f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1['A2'].font = Font(color=MUTED, size=9, name='Consolas')
    ws1['A2'].fill = fill(BG_DARK)
    ws1['A2'].alignment = ctr()

    row = 4
    for label, value in [
        ('Hostname',     srv.get('hostname', 'N/A')),
        ('IP',           srv.get('ip', 'N/A')),
        ('Sistema',      srv.get('os', 'N/A')),
        ('Plataforma',   srv.get('platform', 'linux').upper()),
        ('Fecha Audit.', str(summ.get('audit_date', ''))[:19]),
        ('', ''),
        ('SCORE GLOBAL', f"{summ.get('score', 0)}%"),
        ('Total Checks', summ.get('total', 0)),
        ('PASS',         summ.get('pass', 0)),
        ('FAIL',         summ.get('fail', 0)),
        ('WARN',         summ.get('warn', 0)),
    ]:
        ws1[f'A{row}'] = label
        ws1[f'B{row}'] = value
        ws1[f'A{row}'].font = bfont(CYAN, 9)
        ws1[f'A{row}'].fill = fill(BG_MED)
        ws1[f'A{row}'].alignment = lft()
        ws1[f'B{row}'].font = rfont()
        ws1[f'B{row}'].fill = fill(BG_DARK)
        ws1[f'B{row}'].alignment = lft()
        ws1[f'A{row}'].border = tborder()
        ws1[f'B{row}'].border = tborder()
        ws1.row_dimensions[row].height = 18
        row += 1

    # HOJA 2: CHECKS
    ws2 = wb.create_sheet('Checks')
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'A2'

    headers = ['Estado', 'Categoria', 'Verificacion', 'Descripcion', 'Severidad', 'Detalle']
    widths  = [10, 16, 35, 40, 10, 40]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        col = get_column_letter(i)
        ws2.column_dimensions[col].width = w
        cell = ws2[f'{col}1']
        cell.value = h; cell.font = bfont(CYAN, 9)
        cell.fill = fill(BG_MED); cell.alignment = ctr(); cell.border = tborder()
    ws2.row_dimensions[1].height = 22

    for r, c in enumerate(chks, 2):
        icon = 'PASS' if c['status'] == 'PASS' else 'FAIL' if c['status'] == 'FAIL' else 'WARN'
        row_bg = BG_DARK if r % 2 == 0 else BG_MED
        s_color = RED if c['severity'] == 'ALTA' else YELLOW if c['severity'] == 'MEDIA' else GREEN
        vals = [icon, c['category'], c['name'], c['description'], c['severity'], c.get('detail','')]
        for j, v in enumerate(vals, 1):
            col = get_column_letter(j); cell = ws2[f'{col}{r}']
            cell.value = v; cell.fill = fill(row_bg)
            cell.alignment = lft(); cell.border = tborder(); cell.font = rfont()
        sc = GREEN if c['status']=='PASS' else RED if c['status']=='FAIL' else YELLOW
        ws2[f'A{r}'].font = Font(color=sc, bold=True, size=9, name='Calibri')
        ws2[f'E{r}'].font = Font(color=s_color, bold=True, size=9, name='Calibri')
        ws2.row_dimensions[r].height = 18

    # HOJA 3: CATEGORIAS
    ws3 = wb.create_sheet('Por Categoria')
    ws3.sheet_view.showGridLines = False
    for col, (h, w) in enumerate(zip(['Categoria','Score %','Estado'], [20,12,15]), 1):
        cl = get_column_letter(col); ws3.column_dimensions[cl].width = w
        cell = ws3[f'{cl}1']
        cell.value = h; cell.font = bfont(CYAN, 9)
        cell.fill = fill(BG_MED); cell.alignment = ctr(); cell.border = tborder()
    for r, (cat, pct) in enumerate(cats.items(), 2):
        status = 'OK' if pct >= 80 else 'WARN' if pct >= 60 else 'CRITICO'
        color  = GREEN if pct >= 80 else YELLOW if pct >= 60 else RED
        bg = BG_DARK if r % 2 == 0 else BG_MED
        for col, val in enumerate([cat, f'{pct}%', status], 1):
            cl = get_column_letter(col); cell = ws3[f'{cl}{r}']
            cell.value = val; cell.fill = fill(bg)
            cell.alignment = ctr(); cell.border = tborder()
            cell.font = rfont(color if col > 1 else TEXT)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
